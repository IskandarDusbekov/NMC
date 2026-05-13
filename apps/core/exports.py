"""
Excel export utilities.
openpyxl yordamida .xlsx fayllar yaratadi.
"""
from __future__ import annotations

from io import BytesIO
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from django.http import HttpResponse


# ── style constants ───────────────────────────────────────────────────────
_HEADER_FILL = PatternFill(fill_type='solid', fgColor='0F172A')   # slate-950
_HEADER_FONT = Font(color='FFFFFF', bold=True, size=11)
_TITLE_FONT  = Font(bold=True, size=13)
_ALT_FILL    = PatternFill(fill_type='solid', fgColor='F8FAFC')   # slate-50


def _style_header_row(ws, row_num: int, col_count: int) -> None:
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def _auto_width(ws) -> None:
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)


def build_excel(
    title: str,
    headers: list[str],
    rows: list[list[Any]],
) -> BytesIO:
    """
    Oddiy Excel fayl yaratadi.
    title  — birinchi qatorda katta sarlavha
    headers — ustun nomlari
    rows   — [[val1, val2, ...], ...]
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = _TITLE_FONT
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 24

    # Header row
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=2, column=col_idx, value=header)
    _style_header_row(ws, 2, len(headers))
    ws.row_dimensions[2].height = 20

    # Data rows
    for row_idx, row_data in enumerate(rows, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical='center')
            if row_idx % 2 == 0:
                cell.fill = _ALT_FILL

    _auto_width(ws)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def excel_response(buffer: BytesIO, filename: str) -> HttpResponse:
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
